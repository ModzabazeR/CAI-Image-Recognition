import pdfplumber
import os
import re
import json
import csv
import openpyxl as xl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import pandas as pd
from collections import OrderedDict
import utils

BBL_KEYWORDS = ("Currency THB Date", "By the instruction of", "Beneficiary name :", "Beneficiary Account :",
                "Invoice details as follows (if any)", "Payment Net")
KBANK_KEYWORDS = ("KASIKORNBANK PCL", "On behalf of", "Payment details are as follows")
SCB_KEYWORDS = ("This document is an integral part of Credit Advice", "เรียนท่านเจ้าของบัญชี")
TTB_PATTERN = r"(\d{10})[ ]+([\d.,]+)[ ]+([\d.,]+)[ ]+([\d.,]+)[ ]+([\d.,]+)"
TTB_KEYWORDS = ("0 2299 2176", "0 2299 2572")
BAY_KEYWORDS = ("0-2626-2626", "CMSSupport@krungsri.com")

MAPPING = json.load(open("mapping.json", "r", encoding='utf-8'))

def compile_workbooks(workbooks_path: str, final_filename: str) -> str:
    """
    Compile all workbooks in a given directory into a single workbook.
    """
    try:
        cols = ["A", "B", "C", "D", "E", "F", "G", "H", "P"]
        null_fill = PatternFill(patternType="solid", fgColor="D9D9D9")
        normal_align = Alignment(horizontal="left", vertical="top")
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        header_fill = PatternFill(patternType="solid", fgColor="FFFF99")
        bold = Font(bold=True)
        
        wbs = []
        files = os.listdir(workbooks_path)
        for file in files:
            if not file.startswith("~$") and file.endswith(".xlsx") and "final" not in file and "docjuice" in file:
                wb = xl.load_workbook(os.path.join(workbooks_path, file))
                wbs.append(wb)

        final_wb = xl.Workbook()
        final_ws = final_wb.worksheets[0]
        wb1 = wbs[0]
        ws1 = wb1.worksheets[0]

        # Copy header
        for i in range(1, ws1.max_column + 1):
            final_ws.cell(row=1, column=i).value = ws1.cell(row=1, column=i).value

        # Copy data
        current_row = 2
        for wb in wbs:
            for ws in wb.worksheets:
                start_row = current_row
                mr = ws.max_row
                mc = ws.max_column
                
                for i in range(2, mr + 1):
                    for j in range(1, mc + 1):
                        current_cell = ws.cell(row=i, column=j)
                        final_ws.cell(row=current_row, column=j).value = current_cell.value
                    current_row += 1

                # merge cells
                for col in cols:
                    try:
                        final_ws.merge_cells(f"{col}{start_row}:{col}{current_row - 1}")
                    except ValueError:
                        pass


        # Formatting
        d = final_ws["D2"]
        final_ws.freeze_panes = d
        
        for col_cells in final_ws.columns:
            length = max(len(str(cell.value)) for cell in col_cells)
            final_ws.column_dimensions[col_cells[0].column_letter].width = length

        # styling
        for row in final_ws.iter_rows():
            for cell in row:
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = bold
                if cell.value is None:
                    cell.fill = null_fill
                cell.alignment = normal_align
                cell.border = border

        final_wb.save(os.path.join(workbooks_path, final_filename))
        for file in files:
            if os.path.exists(os.path.join(workbooks_path, file)) and "final" not in file:
                os.remove(os.path.join(workbooks_path, file))
        return "Merge files successfully"
    except Exception as e:
        return f"Error while merging the file: {e}"

# -------- PDF Invoice -------- #
output = ''
class PDFInvoice:
    # Invoice metadata
    pdf = None
    invoice_types = ('Credit Advice Report', 'Pre-Advice Report')
    invoice_extension = ''
    invoice_type = ''
    text = ''

    # Invoice data
    receiver = None
    bank = None
    receiver_account = None
    payment_date = None
    total_after_tax = None
    cheque_id = None
    sender = None
    bank_charge = None

    def close(self) -> None:
        self.pdf.close()

    def get_invoice_info(self) -> None:
        pass

    def extract(self, file: str, mode: str = "records") -> None:
        pass

    def get_entries(self, mode: str = "records") -> dict:
        data = {
            "File Name": os.path.basename(self.file_path),
            "Company (Receive)": self.receiver,
            "Bank": self.bank,
            "Account Bank": self.receiver_account,
            "Date Payment": self.payment_date,
            "Payment Amount": self.total_after_tax,
            "Check No.": self.cheque_id,
            "Payee Name": self.sender,
            "items": self.extract(self.file_path, mode),
            "Fee": self.bank_charge
        }
        return data

    def __init__(self, file_path: str) -> None:
        self.info = None

        self.file_path = file_path
        if file_path.lower().endswith('.pdf'):
            self.invoice_extension = 'pdf'
            self.pdf = pdfplumber.open(file_path)
            self.get_invoice_info()
        else:
            raise Exception('File type not supported')

    def to_txt(self) -> None:
        with open(f'{output}/{os.path.basename(self.file_path)}.txt', 'w', encoding='utf-8') as f:
            for page in self.info:
                f.write(f'[PAGE {page.page_number}]\n') if page.page_number == 1 else f.write(
                    f'\n[PAGE {page.page_number}]\n')
                f.write(utils.correct_words(page.extract_text(), MAPPING))
        print(f'"{os.path.basename(self.file_path)}.txt" written to {output}')

    def to_json(self) -> None:
        data = self.get_entries()
        utils.pretty_save_json(
            f'{output}/{os.path.basename(self.file_path)}.json', data)

        print(f'"{os.path.basename(self.file_path)}.json" written to {output}\n')

    def to_excel(self) -> None:
        global output

        file = self.get_entries(mode="list")

        metadata = {k: v for k, v in file.items() if k != "items" and k !=
                    "Fee"}
        items = file["items"]
        fee = file["Fee"]
        header = [i for i in metadata.keys()] + [i for i in items.keys()] + \
                 ["Fee"]

        # to csv
        with open(f"{output}/{os.path.basename(self.file_path)}.csv", "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            if len(items["Invoice No."]) > 0:
                for i in range(len(items["Invoice No."])):
                    if i == 0:
                        writer.writerow([f"'{metadata[k]}" if metadata[k] is not None else None for k in metadata.keys()] +
                                        [f"'{items[k][i]}" if items[k][i] is not None else None for k in items.keys()] + [f"'{fee}" if fee is not None else None])
                    else:
                        writer.writerow([None for _ in metadata.keys()] +
                                        [f"'{items[k][i]}" if items[k][i] is not None else None for k in items.keys()])
            else:
                writer.writerow([f"'{metadata[k]}" if metadata[k] is not None else None for k in metadata.keys()] +
                                [None for _ in items.keys()] + [f"'{fee}" if fee is not None else None])

        df = pd.read_csv(f"{output}/{os.path.basename(self.file_path)}.csv", encoding="utf-8")

        # remove every cell started by '
        df = df.replace(r'^\'', '', regex=True)

        df.to_excel(f"{output}/docjuice_{os.path.basename(self.file_path)}.xlsx", index=False)
        os.remove(f"{output}/{os.path.basename(self.file_path)}.csv")


class KBANKInvoice(PDFInvoice):

    def parse_row(self, row):
        row = row.split()
        return OrderedDict([
            ("INV.NUMBER", row[0]),
            ("INV.DATE", row[1]),
            ("INV.AMOUNT", row[2]),
            ("VAT AMT", row[3]),
            ("Amt. (Inc. Vat)", ""),
            ("WHT AMT", row[4]),
            ("NET AMOUNT", row[5])
        ])

    def extract(self, path: str, mode="records") -> dict:
        empty_data = {
                    "Invoice No.": [],
                    "Invoice Descriptions": [],
                    "Amt. (Exc. Vat)": [],
                    "Amt Vat.": [],
                    "Amt. (Inc. Vat)": [],
                    "WHT Amt.": [],
                    "Net Amount": []
                }
        pdf = pdfplumber.open(path)
        try:
            p0 = pdf.pages[1]
        except IndexError:
            return empty_data

        if len(pdf.pages) <= 2:        
            text = p0.extract_text()
            core_pat = re.compile(r"NET AMOUNT\n=+\n(.*)\n=+\nTOTAL", re.DOTALL)
            core = re.search(core_pat, text).group(1)
        else:
            core = ""
            for i, page in enumerate(pdf.pages):
                if i == 0:
                    continue
                if i != len(pdf.pages) - 1:
                    text = page.extract_text()
                    core_pat = re.compile(r"NET AMOUNT\n=+\n(.*)\n(TRIAL MODE)*", re.DOTALL)
                    core += f"{re.search(core_pat, text).group(1)}\n"
                else:
                    text = page.extract_text()
                    core_pat = re.compile(r"INVOICE DETAIL\n(.*)\n=+\nTOTAL", re.DOTALL)
                    core += re.search(core_pat, text).group(1)


        if type(core) is not str:
            return empty_data

        core = core.split("\n")

        parsed = [self.parse_row(x) for x in core]
        cols = list(parsed[0].keys())
        data = pd.DataFrame(parsed, columns=cols)
        data = data.drop(columns="INV.DATE")
        data = data.rename(columns={"INV.NUMBER": "Invoice No.", "INV.AMOUNT": "Amt. (Exc. Vat)",
                                    "VAT AMT": "Amt Vat.", "WHT AMT": "WHT Amt.",
                                    "NET AMOUNT": "Net Amount"})
        placeholder = [None for _ in range(len(data))]
        data.insert(1, 'Invoice Descriptions', placeholder)
        data = data.replace([''], [None])
        data_dict = data.to_dict(orient=mode)
        return data_dict

    def get_invoice_info(self) -> None:
        self.info = self.pdf.pages

        for page in self.info:
            self.text += page.extract_text()

        self.text = utils.correct_words(self.text, MAPPING)

        type_match = re.search(r"(Subject *: *)(\w)+", self.text)
        if type_match:
            self.invoice_type = type_match.group(2)

        date_match = re.search(
            r"(Cheque Date|Date) *:? *(\d{2}/\d{2}/\d{4})", self.text)
        if date_match:
            self.payment_date = date_match.group(2)
        else:
            print(
                f'Payment date not found in {os.path.basename(self.file_path)}')

        sender_match = re.search(
            r"(Payer Name|On behalf of) *:? *([\w ()ก-๛.,]+)", self.text)
        if sender_match:
            self.sender = sender_match.group(2)
        else:
            print(
                f'Sender name not found in {os.path.basename(self.file_path)}')

        receiver_match = re.search(r"(To *: *)([\w ()ก-๛.,]+)", self.text)
        if receiver_match:
            self.receiver = receiver_match.group(2)
        else:
            print(
                f'Receiver name not found in {os.path.basename(self.file_path)}')

        total_match = re.search(
            r"(Total Invoice after VAT|Amount) *: *\*+([\d,.]+)", self.text)
        if total_match:
            self.total_after_tax = total_match.group(2)
        else:
            print(
                f'Total after tax not found in {os.path.basename(self.file_path)}')

        bank_charge_match = re.search(
            r"(Benef\.? Charges *: *\*+)([\d,.]+)", self.text)
        if bank_charge_match:
            self.bank_charge = bank_charge_match.group(2)
            self.bank_charge = "0.00" if self.bank_charge == ".00" else self.bank_charge

    def __init__(self, file_path: str) -> None:
        super().__init__(file_path)
        self.bank = 'กสิกรไทย (KBANK)'

class SCBInvoice(PDFInvoice):
    table_settings_p0 = {
    "vertical_strategy": "explicit",
    "horizontal_strategy": "explicit",
    "explicit_vertical_lines": [18, 114, 211, 275, 315, 397, 490, 575],
    "explicit_horizontal_lines": [445, 475, 518, 563, 608, 653, 698, 743]
    }

    table_settings = {
    "vertical_strategy": "explicit",
    "horizontal_strategy": "explicit",
    "explicit_vertical_lines": [18, 114, 211, 275, 315, 397, 490, 575],
    "explicit_horizontal_lines": [18, 50, 95, 140, 183, 226, 269, 312, 355, 397, 439, 482, 525, 568, 611, 654, 697, 740]
    }

    def __init__(self, file_path: str) -> None:
        super().__init__(file_path)
        self.bank = 'ไทยพาณิชย์ (SCB)'

    def get_invoice_info(self) -> None:
        self.info = self.pdf.pages
        for page in self.info:
            self.text += page.extract_text()
        self.text = utils.correct_words(self.text, MAPPING)

        receiver_match = re.search(r'ถึง : ([\w ()ก-๛.,]+) วันที่ (\d{2}-\w{3}-\d{4})', self.text)
        if receiver_match:
            self.receiver = receiver_match.group(1)
        else:
            print(f"Receiver and date not found in {os.path.basename(self.file_path)}")

        payment_date_match = re.search(r'(วันเข้าบัญชี|วันที่บนหน้าเช็ค *): (\d{2}/\d{2}/\d{4})', self.text)
        if payment_date_match:
            self.payment_date = payment_date_match.group(2)
        else:
            print("Payment Date not found")

        total_bank_charge_match = re.search(r'จำนวนเงิน(โอน)* \(บาท\): ([\d,.]+) ค่าธรรมเนียม \(บาท\): ([\d,.]+)', self.text)
        if total_bank_charge_match:
            self.total_after_tax = total_bank_charge_match.group(2)
            self.bank_charge = total_bank_charge_match.group(3)
        else:
            print(f"Total and bank charge not found in {os.path.basename(self.file_path)}")

        sender_match = re.search(r'เรียนท่านเจ้าของบัญชี\n(.+)', self.text)
        if sender_match:
            self.sender = sender_match.group(1)
        else:
            print(f"Sender not found in {os.path.basename(self.file_path)}")

        receiver_account_match = re.search(r'เลขที่บัญชี/หมายเลขพร้อมเพย์: ([\w\d]+\-?[\w\d]+\-?\d+)', self.text)
        if receiver_account_match:
            self.receiver_account = receiver_account_match.group(1)
        else:
            print(f"Receiver Account not found in {os.path.basename(self.file_path)}")

        cheque_id_match = re.search(r'เลขที่เช็ค *: *(\d+)', self.text)
        if cheque_id_match:
            self.cheque_id = cheque_id_match.group(1)
        else:
            print(f"Cheque ID not found in {os.path.basename(self.file_path)}")

    def correct_table(self, table: list) -> list:
        # replace all \n with space
        for i, row in enumerate(table):
            for j, item in enumerate(row):
                table[i][j] = item.replace('\n', ' ')

        # remove blank rows
        table = list(filter(lambda x: x != ['', '', '', '', '', '', ''], table))

        return table

    def extract(self, path: str, mode="records") -> dict:
        pdf = pdfplumber.open(path)
        cols = ["Invoice No.", "Invoice Descriptions", "Invoice Date", "Type of Income", "Invoice Amount", "VAT Amount", "WHT Amount"]
        data = pd.DataFrame(columns=cols)

        for page in pdf.pages:
            if page.page_number == 1:
                table = page.extract_table(self.table_settings_p0)
            else:
                table = page.extract_table(self.table_settings)

            table = self.correct_table(table)
            df = pd.DataFrame(table[1:], columns=cols)
            data = pd.concat([data, df], ignore_index=True)

        # for only detect numbers, but deprecated because the document format is unreliable.
        # for i, row in data[[cols[1]]].iterrows():
        #     data.loc[i, cols[1]] = re.match(r"\d+", row[cols[1]]).group() if re.match(r"\d+", row[cols[1]]) else None

        data = data[["Invoice No.", "Invoice Descriptions", "VAT Amount", "WHT Amount", "Invoice Amount"]]
        data = data.rename(columns={'Invoice No.': 'Invoice No.', 'Invoice Amount': 'Net Amount',
                                    'VAT Amount': 'Amt Vat.', 'WHT Amount': 'WHT Amt.'})

        placeholder = [None for _ in range(len(data))]
        data.insert(2, 'Amt. (Exc. Vat)', placeholder)
        data.insert(4, "Amt. (Inc. Vat)", placeholder)

        data = data.replace([''], [None])
        data_dict = data.to_dict(orient=mode)
        return data_dict

class BBLInvoice(PDFInvoice):
    credit_advice_cols = ["Item No", "Invoice No.", "Date",
                          "Gross Amount", "WHT Amount", "VAT Amount", "Income Type"]
    pre_advice_cols = ["Item No", "Invoice No.",
                       "Date", "Gross Amount", "WHT Amount"]

    def get_invoice_info(self) -> None:
        self.info = self.pdf.pages

        for page in self.info:
            self.text += page.extract_text()

        if all(x in self.text for x in self.credit_advice_cols):
            self.invoice_type = self.invoice_types[0]
        elif all(x in self.text for x in self.pre_advice_cols):
            self.invoice_type = self.invoice_types[1]

        if self.invoice_type not in self.invoice_types:
            raise ValueError(
                f'Invoice type not found in {os.path.basename(self.file_path)}')

        date_match = re.search(
            r'(Payment Date : )(\d{2}-\w{3}-\d{2})', self.text)
        if date_match:
            self.payment_date = date_match.group(2)
        else:
            print(
                f'Payment Date not found in {os.path.basename(self.file_path)}')

        sender_match = re.search(
            r'(By the instruction of : )([\w ()ก-๛.,]+)', self.text)
        if sender_match:
            self.sender = sender_match.group(2)
        else:
            print(f'Sender not found in {os.path.basename(self.file_path)}')

        receiver_match = re.search(
            r'(Beneficiary name : )([\w ()ก-๛.,]+)', self.text)
        if receiver_match:
            self.receiver = receiver_match.group(2)
        else:
            print(f'Receiver not found in {os.path.basename(self.file_path)}')

        receiver_account_match = re.search(
            r'(Beneficiary [Aa]ccount : )(\d+)', self.text)
        if receiver_account_match:
            self.receiver_account = receiver_account_match.group(2)

        total_match = re.search(r'(Payment Net : )([\d,.]+)', self.text)
        if total_match:
            self.total_after_tax = total_match.group(2)

        cheque_id_match = re.search(r'(Cheque No. : )(\d+)', self.text)
        if cheque_id_match:
            self.cheque_id = cheque_id_match.group(2)

        bank_charge_match = re.search(r'Bank Charge.*: *([\d,.]+)', self.text)
        if bank_charge_match:
            self.bank_charge = bank_charge_match.group(1)

    def __init__(self, file_path: str) -> None:
        super().__init__(file_path)
        self.bank = 'กรุงเทพ (BBL)'

    def extract(self, file: str, mode: str = "records") -> dict:
        pdf = pdfplumber.open(file)

        cols = self.credit_advice_cols if self.invoice_type == 'Credit Advice Report' else self.pre_advice_cols

        data = pd.DataFrame(columns=cols)

        for i, page in enumerate(pdf.pages):
            table = page.extract_table()
            if type(table) != list and i == 0:
                return {
                    "Invoice No.": [],
                    "Invoice Descriptions": [],
                    "Amt. (Exc. Vat)": [],
                    "Amt Vat.": [],
                    "Amt. (Inc. Vat)": [],
                    "WHT Amt.": [],
                    "Net Amount": []
                }
            if type(table) != list and i == len(pdf.pages) - 1:
                continue

            # Filter out empty rows
            table = list(filter(lambda a: a != ['', '', '', '', '', '', ''] and a != [
                '', '', '', '', ''], table))

            # replace \n with space in the table
            for i in range(len(table)):
                for j in range(len(table[i])):
                    table[i][j] = table[i][j].replace('\n', ' ')

            for item in table[-1]:
                if "e-WHT" in item:
                    table.pop(-1)
                    break
            df = pd.DataFrame(table, columns=cols)

            data = pd.concat([data, df], ignore_index=True)

        data = data.drop(0)
        data = data[["Invoice No.", "Gross Amount", "WHT Amount"]]
        placeholder = [None for _ in range(len(data))]
        data.insert(1, 'Invoice Descriptions', placeholder)
        data.insert(2, "Amt. (Exc. Vat)", placeholder)
        data.insert(3, "Amt Vat.", placeholder)
        data["Net Amount"] = placeholder
        data = data.drop(data[data["Invoice No."] == "Invoice No."].index)
        # if part of "e-WHT" string in any row, delete the row
        # data = data[~data["Amt. (Inc. Vat)"].str.contains("e-WHT")]
        # data = data[~data["WHT Amt."].str.contains("e-WHT")]

        data = data.rename(columns={"Invoice No.": "Invoice No.",
                                    "Gross Amount": "Amt. (Inc. Vat)", "WHT Amount": "WHT Amt."})
        data = data.replace([''], [None])

        # if all the invoice data are None, return empty dict
        if all(data.isna().all()):
            return {
                    "Invoice No.": [],
                    "Invoice Descriptions": [],
                    "Amt. (Exc. Vat)": [],
                    "Amt Vat.": [],
                    "Amt. (Inc. Vat)": [],
                    "WHT Amt.": [],
                    "Net Amount": []
                }

        data_dict = data.to_dict(orient=mode)
        return data_dict

class TTBInvoice(PDFInvoice):
    def __init__(self, file_path: str) -> None:
        super().__init__(file_path)
        self.bank = 'ทีเอ็มบีธนชาต (TTB)'

    @staticmethod
    def remove_header(table: list) -> list:
        return table[1:]

    @staticmethod
    def correct_table(table: list) -> list:
        for i, row in enumerate(table):
            for j, cell in enumerate(row):
                title_match = re.search(r"([\w ]+) :", cell)
                num_match = re.search(r'THB ([\d,.]+)', cell)
                table[i][j] = title_match.group(1) if title_match else table[i][j]
                table[i][j] = num_match.group(1) if num_match else table[i][j]
        return table
        
    def get_invoice_info(self) -> None:
        p0 = self.pdf.pages[0]
        payment_info, recipient_info, transaction_info = p0.extract_tables()

        payment_info = self.remove_header(payment_info)
        payment_info = self.correct_table(payment_info)

        recipient_info = self.remove_header(recipient_info)
        recipient_info = self.correct_table(recipient_info)

        transaction_info = self.remove_header(transaction_info)
        transaction_info = self.correct_table(transaction_info)

        self.bank_charge = payment_info[1][1]
        self.total_after_tax = payment_info[2][1]
        self.receiver = recipient_info[0][1]
        self.sender = transaction_info[3][1]
        self.payment_date = transaction_info[8][1]
        self.receiver_account = transaction_info[10][1]

    def extract(self, file: str, mode: str = "records") -> dict:
        pdf = self.pdf
        text = ""
        for page in pdf.pages:
            text += page.extract_text()
        text = utils.correct_words(text, MAPPING)

        info = re.findall(r"(\d{10})[ ]+([\d.,]+)[ ]+([\d.,]+)[ ]+([\d.,]+)[ ]+([\d.,]+)", text)
        data = pd.DataFrame(info, columns=["Invoice No.", "Amt. (Exc. Vat)", "Vat Amt.", "WHT Amt.", "Amt. (Inc. Vat)"])
        data = data[["Invoice No.", "Amt. (Exc. Vat)", "Vat Amt.", "Amt. (Inc. Vat)", "WHT Amt."]]
        placeholder = [None for _ in range(len(data))]
        data.insert(1, 'Invoice Descriptions', placeholder)
        data.insert(6, 'Net Amount', data["Amt. (Inc. Vat)"])

        data_dict = data.to_dict(orient=mode)

        return data_dict

class BAYInvoice(PDFInvoice):
    def __init__(self, file_path: str) -> None:
        super().__init__(file_path)
        self.bank = 'กรุงศรีอยุธยา (BAY)'

    def get_invoice_info(self):
        for page in self.pdf.pages:
            self.text += page.extract_text()

        self.text = utils.correct_words(self.text, MAPPING)

        total_after_tax_match = re.search(r"Transaction Amount[ *]+([\d,.]+)", self.text)
        if total_after_tax_match:
            self.total_after_tax = total_after_tax_match.group(1)

        bank_charge_match = re.search(r"Fee Charge[ *]+([\d,.]+)", self.text)
        if bank_charge_match:
            self.bank_charge = bank_charge_match.group(1)
        
        receiver_match = re.search(r"Beneficiary Name[ ]+(.+)", self.text)
        if receiver_match:
            self.receiver = receiver_match.group(1)

        sender_match = re.search(r"By the Order Of[ ]+(.+)", self.text)
        if sender_match:
            self.sender = sender_match.group(1)

        payment_date_match = re.search(r"Effective Date[ ]+(\d+/\d+/\d+)", self.text)
        if payment_date_match:
            self.payment_date = payment_date_match.group(1)

        receiver_account_match = re.search(r"Beneficiary Account[ ]+(\d+)", self.text)
        if receiver_account_match:
            self.receiver_account = receiver_account_match.group(1)

    def extract(self, file: str, mode: str = "records") -> dict:
        data = pd.DataFrame([], columns=["Invoice No.", "Invoice Descriptions", "Amt. (Exc. Vat)", "Vat Amt.", "Amt. (Inc. Vat)", "WHT Amt.", "Net Amount"])
        data_dict = data.to_dict(orient=mode)
        return data_dict